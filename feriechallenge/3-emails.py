'''
Stwórz prosty program, który będzie wysyłał spersonalizowany mailing do wybranych osób. “Bazą danych” jest plik Excela (aby było “ciekawiej” 😉 ) 
lub CSV, zawierający dwie kolumny z nagłówkami: “E-mail” oraz “Imię i nazwisko” (zakładamy, że zawsze w takiej kolejności, 
a imię i nazwisko są oddzielone spacją). 
Do użytkowników należy wysłać maila z tematem “Your image” oraz spersonalizowaną prostą treścią np. “Hi {Imię}! it’s file generated for you”. 
Dodatkowo w załączniku maila znajduje się plik graficzny o nazwie “{Imię}_{Nazwisko}_image.png” (pliki są w zadanej lokalizacji). 

Odpowiednio zabezpiecz program (np. brakujący plik Excela, brakujące dane w Excelu, brak pliku png) oraz 
zabezpiecz przed spamowaniem (np. jeden mail wysyłany co 1 sekundę). Mogą przydać się moduły: smtplib, email, ssl, xlrd, re, os. 

Propozycje rozszerzenia: dodaj opcję wysyłania maili z treścią w HTML oraz walidator poprawności maila (np. używając wyrażeń regularnych - moduł re).
'''
import smtplib
import imghdr
import os
from openpyxl import load_workbook
from email.message import EmailMessage
from time import sleep

''' load xmls and change to list
need excel file emails.xlsx with 2 column data (email, name and surname)
photos with name: “{Imię}_{Nazwisko}_image.jpg”
'''
try:
    wb = load_workbook("emails.xlsx")
    ws = wb['Arkusz1']
except Exception:
    print('The file doesnt exist')


plist = []
for row in ws.iter_rows(values_only=True):
    plist.append(row)

plist.remove(plist[0])

''' Prepare mail 
You have to set gmail first:
1. log in
2. if you have 2 factor auth setup:
go to myaccount.google.com/apppasswords and set the password fot python
else:
go to myaccount.google.com/lesssecureapps
3. place the EMAIL_USER (your email) and EMAIL_PASS (password created above) to env variables
'''

EMAIL_ADDRESS = os.environ.get('EMAIL_USER')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASS')

for i in plist:
    try:
        p = i[1].split(' ')
        msg = EmailMessage()
        msg['Subject'] = f'Your image'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = i[0]
        msg.set_content(f'Hi {p[0]}! its file generated for you')
    except Exception:
        print(f'Some data missin in row {plist.index(i)+2}')

    ''' add image'''

    image = f'{p[0]}_{p[1]}_image.jpg'
    try:
        with open(image, 'rb') as fp:
            img_data = fp.read()
    except Exception:
        print('image doesnt exist or has wrong name')
    msg.add_attachment(img_data, maintype='image', subtype=imghdr.what(None, img_data))

    '''send mail'''

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)

    sleep(1)